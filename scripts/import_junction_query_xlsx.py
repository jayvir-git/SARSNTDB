"""
Build sql/junction_query.sql from Jim Kelley's xlsx files.

Re-run this script if the spreadsheets change, then re-import the SQL.
Group continent/instrument mapping is in GROUP_META below so it can be edited
without touching PHP.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX_DIR = ROOT / "_incoming" / "jim-kelley"
OUT_SQL = ROOT / "sql" / "junction_query.sql"

JUNCTION_LEFT = 5249
JUNCTION_RIGHT = 23191
JUNCTION_SIZE = 17943

# Inferred from group names. Easy to change here and regenerate SQL.
GROUP_META = {
    "SAfr MiSeq": {"location": "South Africa", "continent": "Africa", "instrument": "MiSeq"},
    "Pak iSeq 100": {"location": "Pakistan", "continent": "Asia", "instrument": "iSeq 100"},
    "Est 500": {"location": "Estonia", "continent": "Europe", "instrument": "NextSeq 500"},
    "Port 550": {"location": "Portugal", "continent": "Europe", "instrument": "NextSeq 550"},
    "DE MiSeq": {"location": "Germany", "continent": "Europe", "instrument": "MiSeq"},
    "NJ MiSeq": {"location": "New Jersey", "continent": "N America", "instrument": "MiSeq"},
    "NM MiSeq": {"location": "New Mexico", "continent": "N America", "instrument": "MiSeq"},
    "VA MiSeq": {"location": "Virginia", "continent": "N America", "instrument": "MiSeq"},
    "Arg MiSeq": {"location": "Argentina", "continent": "S America", "instrument": "MiSeq"},
    "Angola MiSeq": {"location": "Angola", "continent": "Africa", "instrument": "MiSeq"},
    "DE NextSeq 550": {"location": "Germany", "continent": "Europe", "instrument": "NextSeq 550"},
    "UK NextSeq 550": {"location": "United Kingdom", "continent": "Europe", "instrument": "NextSeq 550"},
    "UK MiSeq": {"location": "United Kingdom", "continent": "Europe", "instrument": "MiSeq"},
    "NJ-PRJNA708324": {"location": "New Jersey", "continent": "N America", "instrument": None},
    "Port_080823": {"location": "Portugal", "continent": "Europe", "instrument": None},
    "S_Afr-PRJNA636748": {"location": "South Africa", "continent": "Africa", "instrument": None},
    "Africa": {"location": None, "continent": "Africa", "instrument": None},
    "Asia": {"location": None, "continent": "Asia", "instrument": None},
    "Europe": {"location": None, "continent": "Europe", "instrument": None},
    "N America": {"location": None, "continent": "N America", "instrument": None},
    "S America": {"location": None, "continent": "S America", "instrument": None},
}

VARIANT_NORM = {
    "early": ("early", "Early"),
    "alpha": ("alpha", "Alpha"),
    "delta": ("delta", "Delta"),
    "ba.1": ("ba.1", "BA.1"),
    "ba.2": ("ba.2", "BA.2"),
    "ba.4": ("ba.4", "BA.4"),
    "ba.5": ("ba.5", "BA.5"),
    "xbb": ("xbb", "XBB"),
    "omi ba.1": ("ba.1", "BA.1"),
    "omi_ba1": ("ba.1", "BA.1"),
    "omi ba.2": ("ba.2", "BA.2"),
    "other omi": ("other_omi", "Other Omicron"),
    "other_omi": ("other_omi", "Other Omicron"),
    "lambda": ("lambda", "Lambda"),
}


def sql_str(value) -> str:
    if value is None:
        return "NULL"
    return "'" + str(value).replace("\\", "\\\\").replace("'", "''") + "'"


def sql_num(value) -> str:
    if value is None:
        return "NULL"
    return repr(float(value))


def norm_variant(label: str) -> tuple[str, str]:
    key = " ".join(str(label).strip().lower().replace("_", " ").split())
    key = key.replace("omi ", "") if key.startswith("omi ba") else key
    if key in VARIANT_NORM:
        return VARIANT_NORM[key]
    compact = str(label).strip().lower().replace(" ", "_")
    if compact in VARIANT_NORM:
        return VARIANT_NORM[compact]
    display = str(label).strip()
    return (compact, display)


def parse_primer_header(header: str) -> tuple[str, str, str] | None:
    raw = str(header).strip()
    if not raw or "_" not in raw:
        return None
    variant_part, primer_part = raw.rsplit("_", 1)
    if not primer_part.upper().startswith("V"):
        return None
    code, display = norm_variant(variant_part)
    primer = primer_part.upper()
    if primer == "V4.1":
        primer = "V4.1"
    return code, display, primer


def first_worksheet(path: Path):
    wb = load_workbook(path, data_only=True)
    for name in wb.sheetnames:
        ws = wb[name]
        if hasattr(ws, "iter_rows"):
            return ws
    raise RuntimeError(f"No data sheet in {path.name}")


def cell_matrix(ws):
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True):
        rows.append(list(row))
    return rows


def meta_for(name: str) -> dict:
    if name in GROUP_META:
        return GROUP_META[name]
    return {"location": None, "continent": None, "instrument": None}


class Builder:
    def __init__(self) -> None:
        self.datasets: list[dict] = []
        self.groups: list[dict] = []
        self.measures: list[dict] = []
        self._group_key: dict[tuple[int, str], int] = {}
        self._ds_id = 0
        self._g_id = 0
        self._m_id = 0

    def add_dataset(self, code: str, title: str, source_file: str, chart_kind: str, notes: str) -> int:
        self._ds_id += 1
        self.datasets.append(
            {
                "id": self._ds_id,
                "code": code,
                "title": title,
                "source_file": source_file,
                "chart_kind": chart_kind,
                "notes": notes,
            }
        )
        return self._ds_id

    def add_group(self, dataset_id: int, name: str) -> int:
        key = (dataset_id, name)
        if key in self._group_key:
            return self._group_key[key]
        self._g_id += 1
        info = meta_for(name)
        self.groups.append(
            {
                "id": self._g_id,
                "dataset_id": dataset_id,
                "name": name,
                "location_name": info["location"],
                "continent": info["continent"],
                "instrument": info["instrument"],
            }
        )
        self._group_key[key] = self._g_id
        return self._g_id

    def add_measure(
        self,
        dataset_id: int,
        group_id: int,
        variant_code: str,
        variant_label: str,
        primer: str | None,
        source_header: str,
        pct: float,
    ) -> None:
        self._m_id += 1
        self.measures.append(
            {
                "id": self._m_id,
                "dataset_id": dataset_id,
                "group_id": group_id,
                "variant_code": variant_code,
                "variant_label": variant_label,
                "primer": primer,
                "source_header": source_header,
                "pct": pct,
            }
        )


def load_group_variant(b: Builder) -> None:
    path = XLSX_DIR / "LTG_bbmap_vir_072926_17943.xlsx"
    rows = cell_matrix(first_worksheet(path))
    ds = b.add_dataset(
        "group_variant",
        "Percent of samples by group and variant",
        path.name,
        "clustered_bar",
        "Column headers are variants; row headers are groups. -1 means no samples.",
    )
    headers = rows[0]
    for row in rows[1:]:
        name = row[0]
        if not name:
            continue
        gid = b.add_group(ds, str(name).strip())
        for col, header in enumerate(headers):
            if col == 0 or header is None or row[col] is None:
                continue
            code, display = norm_variant(str(header))
            b.add_measure(ds, gid, code, display, None, str(header), float(row[col]))


def load_primer(b: Builder) -> None:
    path = XLSX_DIR / "LTG_bbmap_vir_080726_primer_17943_v3_v4.1.xlsx"
    rows = cell_matrix(first_worksheet(path))
    ds = b.add_dataset(
        "group_primer",
        "Percent of samples by group and variant-primer pair",
        path.name,
        "primer_bar",
        "Alpha_V3 means Alpha with ARTIC V3. Duplicate right-hand V4.1 block is not imported.",
    )
    headers = rows[0]
    for row in rows[1:]:
        name = row[0]
        if not name:
            continue
        gid = b.add_group(ds, str(name).strip())
        seen_headers = set()
        for col, header in enumerate(headers):
            if col == 0 or header is None:
                continue
            parsed = parse_primer_header(str(header))
            if parsed is None:
                continue
            # Skip the duplicated V4.1 block (columns P–V in the workbook).
            if col >= 15:
                continue
            key = (gid, parsed[0], parsed[2])
            if key in seen_headers:
                continue
            seen_headers.add(key)
            if row[col] is None:
                continue
            code, display, primer = parsed
            b.add_measure(ds, gid, code, display, primer, str(header), float(row[col]))


def load_stack_source(b: Builder) -> None:
    path = XLSX_DIR / "var_17943_5249-23191_8_groups_stack-norm.xlsx"
    rows = cell_matrix(first_worksheet(path))
    ds = b.add_dataset(
        "group_variant_8",
        "Eight-group variant percents (stack-norm source)",
        path.name,
        "stacked_bar",
        "Upper table is raw percents. Stacked chart uses percent / number of groups in the query.",
    )
    headers = rows[0]
    for row in rows[1:]:
        name = row[0]
        if not name:
            break
        gid = b.add_group(ds, str(name).strip())
        for col, header in enumerate(headers):
            if col == 0 or header is None or row[col] is None:
                continue
            code, display = norm_variant(str(header))
            b.add_measure(ds, gid, code, display, None, str(header), float(row[col]))


def load_continents(b: Builder) -> None:
    path = XLSX_DIR / "fav-continents-line-17943_NJ1_063025.xlsx"
    rows = cell_matrix(first_worksheet(path))
    ds = b.add_dataset(
        "continent_line",
        "Percent of samples by continent and variant",
        path.name,
        "line",
        "Row headers are continents. Line-chart style from Jim's example workbook.",
    )
    headers = rows[0]
    for row in rows[1:]:
        name = row[0]
        if not name:
            continue
        gid = b.add_group(ds, str(name).strip())
        for col, header in enumerate(headers):
            if col == 0 or header is None or row[col] is None:
                continue
            code, display = norm_variant(str(header))
            b.add_measure(ds, gid, code, display, None, str(header), float(row[col]))


def emit_sql(b: Builder) -> str:
    lines = [
        "-- Junction group / variant / primer query tables.",
        "-- Generated by scripts/import_junction_query_xlsx.py from _incoming/jim-kelley/*.xlsx",
        "-- Reversible: run sql/junction_query_drop.sql",
        "-- Junction prototype: 5249-23191 (inclusive size 17943).",
        "",
        "SET NAMES utf8mb4;",
        "SET FOREIGN_KEY_CHECKS=0;",
        "DROP TABLE IF EXISTS `junction_query_measure`;",
        "DROP TABLE IF EXISTS `junction_query_group`;",
        "DROP TABLE IF EXISTS `junction_query_dataset`;",
        "SET FOREIGN_KEY_CHECKS=1;",
        "",
        "CREATE TABLE `junction_query_dataset` (",
        "  `id` int(10) UNSIGNED NOT NULL AUTO_INCREMENT,",
        "  `code` varchar(64) NOT NULL,",
        "  `title` varchar(255) NOT NULL,",
        "  `source_file` varchar(255) NOT NULL,",
        "  `chart_kind` varchar(32) NOT NULL,",
        "  `junction_left` int(11) NOT NULL,",
        "  `junction_right` int(11) NOT NULL,",
        "  `junction_size` int(11) NOT NULL,",
        "  `notes` text DEFAULT NULL,",
        "  PRIMARY KEY (`id`),",
        "  UNIQUE KEY `uq_jq_dataset_code` (`code`)",
        ") ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;",
        "",
        "CREATE TABLE `junction_query_group` (",
        "  `id` int(10) UNSIGNED NOT NULL AUTO_INCREMENT,",
        "  `dataset_id` int(10) UNSIGNED NOT NULL,",
        "  `name` varchar(128) NOT NULL,",
        "  `location_name` varchar(128) DEFAULT NULL,",
        "  `continent` varchar(32) DEFAULT NULL,",
        "  `instrument` varchar(64) DEFAULT NULL,",
        "  PRIMARY KEY (`id`),",
        "  UNIQUE KEY `uq_jq_group_ds_name` (`dataset_id`,`name`),",
        "  KEY `idx_jq_group_continent` (`continent`),",
        "  KEY `idx_jq_group_instrument` (`instrument`),",
        "  CONSTRAINT `fk_jq_group_dataset` FOREIGN KEY (`dataset_id`) REFERENCES `junction_query_dataset` (`id`) ON DELETE CASCADE",
        ") ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;",
        "",
        "CREATE TABLE `junction_query_measure` (",
        "  `id` int(10) UNSIGNED NOT NULL AUTO_INCREMENT,",
        "  `dataset_id` int(10) UNSIGNED NOT NULL,",
        "  `group_id` int(10) UNSIGNED NOT NULL,",
        "  `variant_code` varchar(64) NOT NULL,",
        "  `variant_label` varchar(64) NOT NULL,",
        "  `primer` varchar(16) DEFAULT NULL,",
        "  `source_header` varchar(64) NOT NULL,",
        "  `pct` decimal(18,10) NOT NULL COMMENT '-1 means no samples',",
        "  PRIMARY KEY (`id`),",
        "  KEY `idx_jq_measure_lookup` (`dataset_id`,`variant_code`,`primer`),",
        "  KEY `idx_jq_measure_group` (`group_id`),",
        "  CONSTRAINT `fk_jq_measure_dataset` FOREIGN KEY (`dataset_id`) REFERENCES `junction_query_dataset` (`id`) ON DELETE CASCADE,",
        "  CONSTRAINT `fk_jq_measure_group` FOREIGN KEY (`group_id`) REFERENCES `junction_query_group` (`id`) ON DELETE CASCADE",
        ") ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;",
        "",
        "INSERT INTO `junction_query_dataset`",
        "  (`id`,`code`,`title`,`source_file`,`chart_kind`,`junction_left`,`junction_right`,`junction_size`,`notes`)",
        "VALUES",
    ]
    ds_rows = []
    for d in b.datasets:
        ds_rows.append(
            "  ({id},{code},{title},{source},{kind},{left},{right},{size},{notes})".format(
                id=d["id"],
                code=sql_str(d["code"]),
                title=sql_str(d["title"]),
                source=sql_str(d["source_file"]),
                kind=sql_str(d["chart_kind"]),
                left=JUNCTION_LEFT,
                right=JUNCTION_RIGHT,
                size=JUNCTION_SIZE,
                notes=sql_str(d["notes"]),
            )
        )
    lines.append(",\n".join(ds_rows) + ";")
    lines.append("")
    lines.append("INSERT INTO `junction_query_group`")
    lines.append("  (`id`,`dataset_id`,`name`,`location_name`,`continent`,`instrument`)")
    lines.append("VALUES")
    g_rows = []
    for g in b.groups:
        g_rows.append(
            "  ({id},{ds},{name},{loc},{cont},{inst})".format(
                id=g["id"],
                ds=g["dataset_id"],
                name=sql_str(g["name"]),
                loc=sql_str(g["location_name"]),
                cont=sql_str(g["continent"]),
                inst=sql_str(g["instrument"]),
            )
        )
    lines.append(",\n".join(g_rows) + ";")
    lines.append("")
    lines.append("INSERT INTO `junction_query_measure`")
    lines.append("  (`id`,`dataset_id`,`group_id`,`variant_code`,`variant_label`,`primer`,`source_header`,`pct`)")
    lines.append("VALUES")
    m_rows = []
    for m in b.measures:
        m_rows.append(
            "  ({id},{ds},{gid},{vcode},{vlab},{primer},{hdr},{pct})".format(
                id=m["id"],
                ds=m["dataset_id"],
                gid=m["group_id"],
                vcode=sql_str(m["variant_code"]),
                vlab=sql_str(m["variant_label"]),
                primer=sql_str(m["primer"]),
                hdr=sql_str(m["source_header"]),
                pct=sql_num(m["pct"]),
            )
        )
    lines.append(",\n".join(m_rows) + ";")
    lines.append("")
    return "\n".join(lines) + "\n"


def main() -> None:
    missing = [
        p.name
        for p in [
            XLSX_DIR / "LTG_bbmap_vir_072926_17943.xlsx",
            XLSX_DIR / "LTG_bbmap_vir_080726_primer_17943_v3_v4.1.xlsx",
            XLSX_DIR / "var_17943_5249-23191_8_groups_stack-norm.xlsx",
            XLSX_DIR / "fav-continents-line-17943_NJ1_063025.xlsx",
        ]
        if not p.exists()
    ]
    if missing:
        raise SystemExit("Missing xlsx files: " + ", ".join(missing))
    b = Builder()
    load_group_variant(b)
    load_primer(b)
    load_stack_source(b)
    load_continents(b)
    OUT_SQL.parent.mkdir(parents=True, exist_ok=True)
    OUT_SQL.write_text(emit_sql(b), encoding="utf-8")
    print(
        "Wrote {out} datasets={d} groups={g} measures={m}".format(
            out=OUT_SQL, d=len(b.datasets), g=len(b.groups), m=len(b.measures)
        )
    )


if __name__ == "__main__":
    main()
